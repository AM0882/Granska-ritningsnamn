import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import re
import tempfile
import pdfplumber
from io import BytesIO

st.title("Jämför ritningsförteckning med PDF-filer")

st.markdown("""
Ladda upp ritningar och ritningsförteckning och jämför.  
I resultatet fås en ritningsförteckning där det står om ritningen är matchad mot en PDF eller inte,  
samt en lista på de ritningar som är med som PDF men inte finns i förteckning.  
OBS ritningar måste ha namn med minst tre olika segment separerade med - eller _ (tex A-1-100 eller W_50.1_1_0100).  
v.1.29
""")

# Upload files
uploaded_pdfs = st.file_uploader("Ladda upp PDF-filer", type=["pdf"], accept_multiple_files=True)
uploaded_reference = st.file_uploader("Ladda upp ritningsförteckning", type=["xlsx", "pdf"])

# Start button
start_processing = st.button("Starta jämförelse")

# Clean text function
def clean_text(text):
    text = str(text).strip().lower()
    return re.sub(r'\.(pdf|docx?|xlsx?|txt|jpg|png|csv)$', '', text)

# Flexible regex for valid drawing/document codes
drawing_pattern = re.compile(
    r'^(?=.*[a-zA-Z])(?=.*\d)[a-zA-Z0-9]+([-_][a-zA-Z0-9]+){2,}$',
    re.IGNORECASE
)

# Regex for extracting codes from PDF (handles both formats)
drawing_number_pattern = re.compile(r'\b[0-9A-Za-z]+(?:[-_][0-9A-Za-z]+){2,}\b')

if start_processing and uploaded_pdfs and uploaded_reference:
    with st.spinner("Bearbetar filer..."):
        try:
            progress = st.progress(0)
            total_steps = 6

            # Step 1: Extract and clean PDF filenames
            pdf_names = [pdf.name for pdf in uploaded_pdfs]
            cleaned_pdf_names = [clean_text(name) for name in pdf_names]
            progress.progress(1 / total_steps)

            # Step 2: Read and clean reference text
            reference_texts = []

            if uploaded_reference.name.lower().endswith(".xlsx"):
                excel_bytes = BytesIO(uploaded_reference.read())
                df_ref = pd.read_excel(excel_bytes, header=None, engine="openpyxl")
                reference_texts = df_ref.astype(str).stack().map(clean_text).tolist()

            elif uploaded_reference.name.lower().endswith(".pdf"):
                pdf_bytes = BytesIO(uploaded_reference.read())
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
                    temp_pdf.write(pdf_bytes.read())
                    temp_pdf.flush()
                    with pdfplumber.open(temp_pdf.name) as pdf:
                        for page in pdf.pages:
                            page_text = page.extract_text()
                            if page_text:
                                matches = drawing_number_pattern.findall(page_text)
                                reference_texts.extend([clean_text(m) for m in matches])

            progress.progress(2 / total_steps)

            # Step 3: Filter reference texts using regex, exclude dates
            filtered_reference_texts = [
                ref for ref in reference_texts
                if drawing_pattern.match(ref)
                and not re.match(r'^202\d', ref)  # Exclude dates starting with 202
            ]

            # Prepare preview DataFrame for reference drawings
            preview_data = []
            for ref in filtered_reference_texts:
                match_status = "Matchad" if ref in cleaned_pdf_names else "Ej matchad"
                preview_data.append({"Referensnamn": ref, "Matchstatus": match_status})

            preview_df = pd.DataFrame(preview_data)

            # Show preview table for reference drawings
            st.subheader("Förhandsgranskning av ritningsförteckning")
            st.dataframe(preview_df)

            # Prepare unmatched PDF names preview
            unmatched_cleaned = [name for name in cleaned_pdf_names if name not in filtered_reference_texts]
            unmatched_df = pd.DataFrame(unmatched_cleaned, columns=["Omatchade PDF-namn"])

            st.subheader("Förhandsgranskning av omatchade PDF-filer")
            st.dataframe(unmatched_df)

            # Step 4: Create Excel with match status and highlight matches
            wb = Workbook()
            ws = wb.active
            ws.title = "Ritningsförteckning"
            ws.append(["Referensnamn", "Matchstatus"])
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row in preview_data:
                ws.append([row["Referensnamn"], row["Matchstatus"]])
                if row["Matchstatus"] == "Matchad":
                    for cell in ws[ws.max_row]:
                        cell.fill = fill

            progress.progress(3 / total_steps)

            # Step 5: Save highlighted Excel file
            result_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(result_file.name)
            progress.progress(4 / total_steps)

            # Step 6: Save unmatched PDF names
            unmatched_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            unmatched_df.to_excel(unmatched_file.name, index=False)
            progress.progress(1.0)

            # Store files in session state for persistent download buttons
            st.session_state["result_file"] = result_file.name
            st.session_state["unmatched_file"] = unmatched_file.name

            st.success("Bearbetning klar!")

        except Exception as e:
            st.error(f"Ett fel uppstod: {e}")

# Show download buttons if files exist in session state
if "result_file" in st.session_state and "unmatched_file" in st.session_state:
    st.download_button(
        "Ladda ner ritningsförteckning med markering",
        data=open(st.session_state["result_file"], "rb").read(),
        file_name="ritningsförteckning_markering.xlsx"
    )
    st.download_button(
        "Ladda ner lista på omatchade PDFer",
        data=open(st.session_state["unmatched_file"], "rb").read(),
        file_name="omatchade_ritningar.xlsx"
    )
